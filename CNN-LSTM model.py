from tensorflow.keras.preprocessing.image import ImageDataGenerator
from tensorflow.keras.layers import Dense, BatchNormalization, Dropout, Bidirectional, Input, Flatten,\
    TimeDistributed, LSTM, RNN, GRU, GlobalAveragePooling2D, SimpleRNN
from tensorflow.keras.applications.resnet import ResNet50
from tensorflow.keras.applications.vgg19 import VGG19
from tensorflow.keras.applications.densenet import DenseNet121
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix, multilabel_confusion_matrix
from keras.applications.resnet import preprocess_input
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import OneHotEncoder
from random import shuffle
from keras import regularizers
import tensorflow as tf
import os
import matplotlib.pyplot as plt
import numpy as np
import cv2
import time
import itertools
import openpyxl
import pandas as pd

start = time.time()

tf.compat.v1.disable_eager_execution()

# Error blocking
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

# Motion_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
#                29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50, 51, 52, 53, 54,
#                55, 56, 57, 58, 59, 60, 61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80]
Motion_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
               29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40]
Validation_subject = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27,
                      28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40]
Test_subject = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
                29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40]

Validation_repetition = [2]
Test_repetition = [5]

Delimeter = ["Subject_", "_Motion_", "_trial_", "_ch_12"]  # EMG
FILE_NAME_START = "Subject"
FILE_NAME_END = ".png"

MODEL_NAME = ResNet50
Motion_lenght = 2
NUM_OF_OUTPUT = len(Motion_list)
Window_size = 800
Overlap_size = 600
IMG_SIZE = 224                      # reshaped input shape (img_size X img_size)
DEPTH = 3                           # reshaped input shape (img_size X img_size X Depth)
BATCH_SIZE = 8                    # batch size
TIME_STEP = 5
INITIAL_EPOCHS = 30                # Number of epoch
Base_learning_rate = 0.0001         # Adam : 0.0001
DROP_OUT_RATE = 0.5                 # Default 0.5
KERNEL_INITIALIZER = 'he_uniform'   # he normal glorot uniform
REGULARIZER = regularizers.l2(0.0001)
ACTIVATION_NAME = 'relu'
OPTIMIZER = tf.keras.optimizers.Adamax(lr=Base_learning_rate)
LOSS = "categorical_crossentropy"

# TNS STS
data_location = "Z:/Nina2 DB/Bandpass(50-500)_folder/1_Avg_slope_Transient/1_Motion40_Subject40_TNS_gray_img/"
# "Z:/Nina2 DB/Bandpass(50-500)_folder/1_Avg_slope_Transient/1_Motion40_Subject40_TNS_gray_img"
# "D:/Whole_TNS_STS/Normal/"
# "D:/Test/TNS"
# data_sub_location = "/"
Classification_label_name = 'Motion_'

# TODO : SAVE EXCEL
Excel_save_location = 'Z:/Nina2 DB/Confusion_matrix/'
First_sheet_Title = 'Predicted Data'
Second_sheet_Title = 'Confusion Matrix'
File_Name_for_segment = "_segment_"
Excel_save_extension = 'M40_S40_CNN-LSTM_TRANSIENT_W400_0200.xlsx'


class InputData:  # input data initialization class
    """""
    After chaning the train img to np.array format, labeling according to Motions
    """""
    train_input = []
    train_label = []
    train_name = []
    out_train_input = []
    out_train_label = []
    out_train_name = []

    test_input = []
    test_label = []
    test_name = []
    out_test_input = []
    out_test_label = []
    out_test_name = []

    validation_input = []
    validation_label = []
    validation_name = []
    out_val_input = []
    out_val_label = []
    out_val_name = []

    tmp1 = []
    tmp2 = []
    tmp3 = []

    def input_image_initialization(self, folder_list, data_gen):
        """""
        Labeling input and label by Motions and Save
        """""
        # Onehot encoder
        label_encoder = LabelEncoder()
        integer_encoded = label_encoder.fit_transform(folder_list)
        onehot_encoder = OneHotEncoder(sparse=False)
        integer_encoded = integer_encoded.reshape(len(integer_encoded), 1)
        onehot_encoded = onehot_encoder.fit_transform(integer_encoded)

        # Data Initialization according to Motions
        for img in data_gen.filenames:
            try:
                img_path = os.path.join(data_gen.directory, img)  # .directory: train_dir || img = filenames
                name = img_path[img_path.find(Classification_label_name):
                                img_path.find(Classification_label_name) + len(Classification_label_name)
                                + Motion_lenght]
                file_name = img_path[img_path.find(FILE_NAME_START):img_path.find(FILE_NAME_END)]

                for list_counter in range(len(Motion_list)):
                    if int(name[len(name)-2:len(name)]) == int(Motion_list[list_counter]):
                        label_number = list_counter
                        self.img_append_inter_repetition(path=img_path, path_name=name, file_name=file_name,
                                                         split_window_size=Window_size, overlap_size=Overlap_size,
                                                         onehot=onehot_encoded, label_num=label_number)

            except Exception as e:
                print(img, e)

        # Merge input and label according to Motions
        self.tmp1 = [[x, y, z] for x, y, z in zip(self.train_input, self.train_label, self.train_name)]
        # shuffle(self.tmp1)
        self.out_train_input = [arr[0] for arr in self.tmp1]
        self.out_train_label = [arr[1] for arr in self.tmp1]
        self.out_train_name = [arr[2] for arr in self.tmp1]

        self.tmp1 = []
        self.train_input = []
        self.train_label = []
        self.train_name = []

        self.tmp2 = [[x, y, z] for x, y, z in zip(self.validation_input, self.validation_label, self.validation_name)]
        # shuffle(self.tmp2)
        self.out_val_input = [arr[0] for arr in self.tmp2]
        self.out_val_label = [arr[1] for arr in self.tmp2]
        self.out_val_name = [arr[2] for arr in self.tmp2]

        self.tmp2 = []
        self.validation_input = []
        self.validation_label = []
        self.validation_name = []

        self.tmp3 = [[x, y, z] for x, y, z in zip(self.test_input, self.test_label, self.test_name)]
        # shuffle(self.tmp3)
        self.out_test_input = [arr[0] for arr in self.tmp3]
        self.out_test_label = [arr[1] for arr in self.tmp3]
        self.out_test_name = [arr[2] for arr in self.tmp3]

        self.tmp3 = []
        self.test_input = []
        self.test_label = []
        self.test_name = []

        num_train_index = len(self.out_train_input)
        num_val_index = len(self.out_val_input)
        num_test_index = len(self.out_test_input)

        delete_train_index = num_train_index % 30
        delete_val_index = num_val_index % 30
        delete_test_index = num_test_index % 30

        print(len(self.out_train_input), len(self.out_train_label))
        print(len(self.out_val_input), len(self.out_val_label))
        print(len(self.out_test_input), len(self.out_test_label))

        del self.out_train_input[-delete_train_index:]
        del self.out_val_input[-delete_val_index:]
        del self.out_test_input[-delete_test_index:]

        del self.out_train_label[-delete_train_index:]
        del self.out_val_label[-delete_val_index:]
        del self.out_test_label[-delete_test_index:]

        print(len(self.out_train_input), len(self.out_train_label))
        print(len(self.out_val_input), len(self.out_val_label))
        print(len(self.out_test_input), len(self.out_test_label))

        # TODO 5: CHANGE TRAIN_INPUT, LABEL DATA SHAPE AND TYPE
        # np.array is 1-dimensional array therefore, to make the original 224x224x3 form
        self.out_train_input = np.reshape(self.out_train_input, (-1, TIME_STEP, IMG_SIZE, IMG_SIZE, DEPTH))
        self.out_train_label = np.reshape(self.out_train_label, (-1, TIME_STEP, NUM_OF_OUTPUT))
        self.out_train_input = np.array(self.out_train_input).astype(np.uint8)
        self.out_train_label = np.array(self.out_train_label).astype(np.uint8)

        self.out_val_input = np.reshape(self.out_val_input, (-1, TIME_STEP, IMG_SIZE, IMG_SIZE, DEPTH))
        self.out_val_label = np.reshape(self.out_val_label, (-1, TIME_STEP, NUM_OF_OUTPUT))
        self.out_val_input = np.array(self.out_val_input).astype(np.uint8)
        self.out_val_label = np.array(self.out_val_label).astype(np.uint8)

        self.out_test_input = np.reshape(self.out_test_input, (-1, TIME_STEP, IMG_SIZE, IMG_SIZE, DEPTH))
        self.out_test_label = np.reshape(self.out_test_label, (-1, TIME_STEP, NUM_OF_OUTPUT))
        self.out_test_input = np.array(self.out_test_input).astype(np.uint8)
        self.out_test_label = np.array(self.out_test_label).astype(np.uint8)

    def img_append_inter_repetition(self, path, path_name, file_name, split_window_size,
                                    overlap_size, onehot, label_num):

        subject_number = int(file_name[file_name.find(Delimeter[0])+len(Delimeter[0]):file_name.find(Delimeter[1])])

        repetition_number = int(file_name[file_name.find(Delimeter[2]) + len(Delimeter[2]):
                                          file_name.find(Delimeter[3])])

        IMG = cv2.imread(path, cv2.IMREAD_COLOR)  # Read images

        image_height = IMG.shape[0]
        image_width = IMG.shape[1]

        # count the maximum number of segmantation. window: 400, overlap: 40
        # ex) 1000-400/400-360 + 1
        number_of_segmentation = ((image_width - split_window_size) // (split_window_size - overlap_size)) + 1

        if image_width >= split_window_size:
            for segment in range(number_of_segmentation):
                start_point = segment * (split_window_size - overlap_size)  # get starting point of spilt image
                split_img = IMG[0:image_height, start_point:(start_point + split_window_size)]  # split Image
                split_img = cv2.resize(split_img, dsize=(IMG_SIZE, IMG_SIZE),
                                       interpolation=cv2.INTER_AREA)  # Img resize

                if (repetition_number in Test_repetition) and (subject_number in Test_subject):
                    # image to array
                    self.test_input.append([np.array(split_img)])
                    # labeling according to Motions (same label for all segment)
                    self.test_label.append([np.array(onehot[label_num])])
                    self.test_name.append(file_name + File_Name_for_segment + str(segment))  # change the name
                    print(path_name, ", Trial :", repetition_number, ', Test data Convert Time: ',
                          time.time() - start, ', Test Count: ', len(self.test_label), ', Label: ',
                          self.test_label[len(self.test_label) - 1], ", Test")

                elif (repetition_number in Validation_repetition) and (subject_number in Validation_subject):
                    self.validation_input.append([np.array(split_img)])
                    self.validation_label.append([np.array(onehot[label_num])])
                    self.validation_name.append(file_name + File_Name_for_segment + str(segment))  # change the name
                    print(path_name, ", Trial :", repetition_number, ', Validation data Convert Time: ',
                          time.time() - start, ', Validation Count: ', len(self.validation_label), ', Label: ',
                          self.validation_label[len(self.validation_label) - 1], ", Validation")

                else:
                    self.train_input.append([np.array(split_img)])
                    self.train_label.append([np.array(onehot[label_num])])
                    self.train_name.append(file_name + File_Name_for_segment + str(segment))  # change the name
                    print(path_name, ", Trial :", repetition_number, ', Training data Convert Time: ',
                          time.time() - start, ', Training Count: ', len(self.train_label), ', Label: ',
                          self.train_label[len(self.train_label) - 1], ", Train")


def visualization_result(x):
    acc = x.history['accuracy']
    val_acc = x.history['val_accuracy']
    loss = x.history['loss']
    val_loss = x.history['val_loss']

    epochs_range = range(1, len(acc) + 1)

    plt.figure(figsize=(8, 8))
    plt.subplot(1, 2, 1)
    plt.plot(epochs_range, acc, label='Training Accuracy')
    plt.plot(epochs_range, val_acc, label='Validation Accuracy')
    plt.legend(loc='lower right')
    plt.title('Training and Validation Accuracy')

    plt.subplot(1, 2, 2)
    plt.plot(epochs_range, loss, label='Training Loss')
    plt.plot(epochs_range, val_loss, label='Validation Loss')
    plt.legend(loc='upper right')
    plt.title('Training and Validation Loss')

    plt.savefig('The_result.png')
    plt.clf()


def plot_confusion_matrix(cm, classes, normalize=False, title='Confusion matrix', cmap=plt.cm.Blues):
    """
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting 'normalize=True'.
    """
    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        cm = np.round(cm, 2)
        print("Normalized confusion matrix")
        title_sub = "Normalized"
    else:
        print("Confusion matrix, without normalizaiton")
        title_sub = "with out Normalized"

    plt.imshow(cm, interpolation='nearest', cmap=cmap)
    plt.title(title)
    plt.colorbar()
    tick_marks = np.arange(len(classes))
    plt.xticks(tick_marks, classes, rotation=45)
    plt.yticks(tick_marks, classes)

    # fmt = '0:.1f' if normalize else 'd'  # 수정
    thresh = cm.max() / 2.
    for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
        plt.text(j, i, cm[i, j], horizontalalignment="center",
                 color="white" if cm[i, j] > thresh else "black")

    plt.tight_layout()
    plt.ylabel('True label')
    plt.xlabel('Predicted label')
    plt.savefig(title+title_sub+".png")
    plt.clf()


def saving_result_excel(excel_save_location, first_sheet_title, classification_label_name,
                        motion_list, y_test_name, y_test, y_predict):

    today_time = time.localtime(time.time())
    today_year = str(today_time.tm_year)
    today_month = '%02d' % today_time.tm_mon
    today_day = '%02d' % today_time.tm_mday
    today = today_year + today_month + today_day

    excel_save_name = excel_save_location + today + Excel_save_extension

    wb = openpyxl.Workbook()
    ws1 = wb.active

    ws1.title = first_sheet_title

    first_line = []
    first_line.append('File_name')
    first_line.append('True_value')
    first_line.append('Estimation_value')
    first_line.append('Match')

    for element in range(len(motion_list)):
        classification_cell = '%02d' % motion_list[element]
        first_line.append(classification_label_name+'_' + classification_cell)
    ws1.append(first_line)

    for row in range(len(y_test_name)):
        ws1.cell(row=row + 2, column=1).value = y_test_name[row]
        ws1.cell(row=row + 2, column=2).value = motion_list[np.argmax(y_test[row], 0)]
        ws1.cell(row=row + 2, column=3).value = motion_list[np.argmax(y_predict[row], 0)]
        ws1.cell(row=row + 2, column=4).value = np.equal(int(np.argmax(y_test[row], 0)),
                                                         int(np.argmax(y_predict[row], 0)))

    for row in range(len(y_test_name)):
        for column in range(len(motion_list)):
            ws1.cell(row=row + 2, column=5+column).value = y_predict[row][column]

    wb.save(excel_save_name)


# def saving_result_excel(excel_save_location, first_sheet_title, second_sheet_title, classification_label_name,
#                         motion_list, cnf_matrix, y_test_name, y_test, y_predict):
#
#     today_time = time.localtime(time.time())
#     today_year = str(today_time.tm_year)
#     today_month = '%02d' % today_time.tm_mon
#     today_day = '%02d' % today_time.tm_mday
#     today = today_year + today_month + today_day
#
#     excel_save_name = excel_save_location + today + Excel_save_extension
#
#     wb = openpyxl.Workbook()
#     ws1 = wb.active
#
#     ws1.title = first_sheet_title
#     # ws2 = wb.create_sheet(second_sheet_title, 1)
#
#     first_line = []
#     first_line.append('File_name')
#     first_line.append('True_value')
#     first_line.append('Estimation_value')
#     first_line.append('Match')
#
#     for element in range(len(motion_list)):
#         classification_cell = '%02d' % motion_list[element]
#         first_line.append(classification_label_name+'_' + classification_cell)
#     ws1.append(first_line)
#
#     for row in range(len(y_test_name)):
#         ws1.cell(row=row + 2, column=1).value = y_test_name[row]
#         ws1.cell(row=row + 2, column=2).value = motion_list[np.argmax(y_test[row], 0)]
#         ws1.cell(row=row + 2, column=3).value = motion_list[np.argmax(y_predict[row], 0)]
#         ws1.cell(row=row + 2, column=4).value = np.equal(int(np.argmax(y_test[row], 0)),
#                                                          int(np.argmax(y_predict[row], 0)))
#
#     for row in range(len(y_test_name)):
#         for column in range(len(motion_list)):
#             ws1.cell(row=row + 2, column=5+column).value = y_predict[row][column]
#
#     # ws2.merge_cells(start_row=1, end_row=1, start_column=2, end_column=len(motion_list)+2)
#     # ws2.cell(row=1, column=2).value = 'Predicted Label'
#     # ws2.merge_cells(start_row=2, end_row=len(motion_list) + 2, start_column=1, end_column=1)
#     # ws2.cell(row=2, column=1).value = 'True Label'
#     #
#     # for column in range(len(motion_list)):
#     #     classification_cell = '%02d' % motion_list[column]
#     #     ws2.cell(row=2, column=column + 3).value = classification_label_name + '_' + classification_cell
#     #
#     # for row in range(len(motion_list)):
#     #     classification_cell = '%02d' % motion_list[row]
#     #     ws2.cell(row=row + 3, column=2).value = classification_label_name + '_' + classification_cell
#     #
#     # for row in range(len(cnf_matrix)):
#     #     for column in range(len(cnf_matrix[0])):
#     #         ws2.cell(row=row + 3, column=column + 3).value = cnf_matrix[row, column]
#
#     wb.save(excel_save_name)


def create_model(img_rows, img_colms, depth, model_name, kernel_init, regular_init, time_step,
                 drop_out, act_name, num_out):

    model_input = tf.keras.layers.Input(shape=(img_rows, img_colms, depth))

    cnnmodel = model_name(include_top=False, weights='imagenet', input_tensor=model_input)
    for layer in cnnmodel.layers:
        layer.trainable = True

    crnn_model = tf.keras.Sequential()
    crnn_model.add(TimeDistributed(cnnmodel, input_shape=(time_step, img_rows, img_colms, depth), name="CNN_LSTM_INPUT"))
    crnn_model.add(TimeDistributed(Dense(256, activation=act_name, kernel_initializer=kernel_init,
                                         kernel_regularizer=regular_init), name="CNN_LSTM_LAYERS_1"))
    crnn_model.add(BatchNormalization(name="CNN_LSTM_BN_1"))
    crnn_model.add(Dropout(drop_out, name="CNN_LSTM_DO_1"))

    crnn_model.add(TimeDistributed(GlobalAveragePooling2D(name="CNN_LSTM_GAP")))

    # acv: tanh -> sigmoid, loss
    crnn_model.add(Bidirectional(LSTM(128, return_sequences=True, kernel_initializer="glorot_normal",
                                      recurrent_initializer="random_normal", bias_initializer="he_uniform",
                                      kernel_regularizer=regular_init, activity_regularizer=regular_init,
                                      activation="tanh", recurrent_activation="hard_sigmoid"), name="LSTM_1"))
    crnn_model.add(Bidirectional(LSTM(128, return_sequences=True, kernel_initializer="glorot_normal",
                                      recurrent_initializer="random_normal", bias_initializer="he_uniform",
                                      kernel_regularizer=regular_init, activity_regularizer=regular_init,
                                      activation="tanh", recurrent_activation="hard_sigmoid"), name="LSTM_2"))

    crnn_model.add(TimeDistributed(Dense(128, activation=act_name, kernel_initializer=kernel_init,
                                         kernel_regularizer=regular_init), name="CNN_LSTM_LAYERS_2"))
    crnn_model.add(BatchNormalization(name="CNN_LSTM_BN_2"))
    crnn_model.add(Dropout(drop_out, name="CNN_LSTM_DO_2"))
    crnn_model.add(TimeDistributed(Dense(64, activation=act_name, kernel_initializer=kernel_init,
                                         kernel_regularizer=regular_init), name="CNN_LSTM_LAYERS_3"))
    crnn_model.add(BatchNormalization(name="CNN_LSTM_BN_3"))
    crnn_model.add(Dropout(drop_out, name="CNN_LSTM_DO_3"))
    crnn_model.add(TimeDistributed(Dense(num_out, activation='softmax'), name="CNN_LSTM_LAYERS_4"))

    return crnn_model


gpus = tf.config.experimental.list_physical_devices('GPU')

if gpus:
    try:
        for gpu in gpus:
            tf.config.experimental.set_memory_growth(gpu, True)
        logical_gpus = tf.config.experimental.list_logical_devices('GPU')
        print(len(gpus), "Physical GPUs,", len(logical_gpus), "Logical GPUs")
    except RuntimeError as e:
        # Memory growth must be set before GPUs have been initialized
        print(e)

# TODO 1: READ TRAIN DATA
# TNS STS
data_dir = data_location
data_folder_list = np.array(os.listdir(data_dir))
data_motion_dir = os.path.join(data_dir, Classification_label_name)

# TODO 2: DATA AUGMENTATION
data_image_generator = ImageDataGenerator(preprocessing_function=preprocess_input)
data_data_gen = data_image_generator.flow_from_directory(directory=data_dir)

# TODO 3: Input img Initialization(labeling, img to array)
input_data = InputData()  # Call class
input_data.input_image_initialization(data_gen=data_data_gen, folder_list=data_folder_list)

# TODO 4: MODEL
print("create model")
pre_trained_model = create_model(img_rows=IMG_SIZE, img_colms=IMG_SIZE, depth=DEPTH, model_name=MODEL_NAME,
                                 kernel_init=KERNEL_INITIALIZER, regular_init=REGULARIZER, drop_out=DROP_OUT_RATE,
                                 act_name=ACTIVATION_NAME, num_out=NUM_OF_OUTPUT, time_step=TIME_STEP)

# TODO 5: COMPILE THE MODEL(LOSS AND OPTIMIZER)
# Define Loss function, Optimizer
pre_trained_model.compile(optimizer=OPTIMIZER, loss=LOSS, metrics=['accuracy'])
pre_trained_model.summary()

# TODO 6: TRAIN THE MODEL
# fit: Our entire training set can fit into RAM, There is no data augmentation going on.
X_train = input_data.out_train_input
Y_train = input_data.out_train_label

X_val = input_data.out_val_input
Y_Val = input_data.out_val_label

X_test = input_data.out_test_input
Y_test = input_data.out_test_label

Y_test_name = input_data.out_test_name


# # TODO: Early stopping
# early_stop = tf.keras.callbacks.EarlyStopping(monitor='val_loss', mode='min', verbose=1, patience=10)

# TODO 7: TRAIN THE MODEL
history = pre_trained_model.fit(x=X_train, y=Y_train, batch_size=BATCH_SIZE, epochs=INITIAL_EPOCHS,
                                validation_data=(X_val, Y_Val), verbose=2)
                                # , callbacks=[early_stop])

pre_trained_model.evaluate(X_test, Y_test, batch_size=BATCH_SIZE, verbose=2)

# pre_trained_model.save("Z:/개인 폴더/김수열/시나리오/SCIE/accuracy/Model/tns_model_labeling80.h5")

# # TODO 7: VISUALIZE TRAINING RESULTS
# visualization_result(history)

# TODO 8: CONFUSION MATRIX
Y_predict = pre_trained_model.predict(X_test)

# TODO: reshape
Y_test_name_reshape = []
Y_test_name_reshape = Y_test_name[0::5]
del Y_test_name_reshape[-3:]

Y_test_reshape = Y_test.mean(axis=1)
Y_predict_reshape = Y_predict.mean(axis=1)

# CONFUSION_MATRIX = confusion_matrix(np.argmax(Y_test, 1), np.argmax(Y_predict, 1))

today_time = time.localtime(time.time())
today_year = str(today_time.tm_year)
today_month = '%02d' % today_time.tm_mon
today_day = '%02d' % today_time.tm_mday
today = today_year + today_month + today_day

excel_save_name = Excel_save_location + today + Excel_save_extension

wb = openpyxl.Workbook()
ws1 = wb.active

ws1.title = First_sheet_Title

first_line = []
first_line.append('File_name')
first_line.append('True_value')
first_line.append('Estimation_value')
first_line.append('Match')

for element in range(len(Motion_list)):
    classification_cell = '%02d' % Motion_list[element]
    first_line.append(Classification_label_name + '_' + classification_cell)
ws1.append(first_line)

for row in range(len(Y_test_name_reshape)):
    ws1.cell(row=row + 2, column=1).value = Y_test_name_reshape[row]
    ws1.cell(row=row + 2, column=2).value = Motion_list[np.argmax(Y_test_reshape[row], 0)]
    ws1.cell(row=row + 2, column=3).value = Motion_list[np.argmax(Y_predict_reshape[row], 0)]
    ws1.cell(row=row + 2, column=4).value = np.equal(int(np.argmax(Y_test_reshape[row], 0)),
                                                     int(np.argmax(Y_predict_reshape[row], 0)))

for row in range(len(Y_test_name_reshape)):
    for column in range(len(Motion_list)):
        ws1.cell(row=row + 2, column=5 + column).value = Y_predict_reshape[row][column]

wb.save(excel_save_name)